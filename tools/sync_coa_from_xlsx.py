# -*- coding: utf-8 -*-
"""Build Full_ChartOfAccounts_Migration.sql from Accounts.xlsx (Desktop path configurable)."""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    raise

XLSX_DEFAULT = Path.home() / "Desktop" / "Accounts.xlsx"
OUT_SQL = Path(__file__).resolve().parents[1] / "Full_ChartOfAccounts_Migration.sql"


def norm_code(v) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return str(int(v))
    s = str(v).strip()
    if not s or not re.match(r"^\d+$", s):
        return None
    return s


def sql_str(s: str | None) -> str:
    if s is None or str(s).strip() == "":
        return "NULL"
    t = str(s).strip().replace("'", "''")
    return f"'{t}'"


def load_accounts(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    out: list[dict] = []
    for row in rows:
        if not row or row[0] is None:
            continue
        if row[0] == "الرمز":
            continue
        if isinstance(row[0], str) and "شجرة" in row[0]:
            continue
        code = norm_code(row[0])
        if not code:
            continue
        name = (row[1] or "").strip() if len(row) > 1 else ""
        desc = row[3] if len(row) > 3 else None
        if desc is not None and isinstance(desc, str):
            desc = desc.strip() or None
        posting = row[5] if len(row) > 5 else None
        if isinstance(posting, str):
            posting = posting.strip()
        allow = 1 if str(posting).lower() == "yes" else 0

        # ملف Excel يكرر 110103 لـ «نقدية الحسابات» و«العجز والزيادة» — الثاني يصبح 110104
        if code == "110103" and ("عجز" in name or "زيادة" in name):
            code = "110104"

        out.append(
            {
                "code": code,
                "name": name,
                "desc": desc,
                "allow": allow,
            }
        )

    by_code: dict[str, dict] = {}
    for a in out:
        by_code[a["code"]] = a
    accounts = list(by_code.values())

    return accounts


def type_nature(code: str) -> tuple[int, int]:
    c = code[0]
    if c == "1":
        return 1, 1
    if c == "2":
        return 2, 2
    if c == "3":
        return 3, 2
    if c == "4":
        return 4, 2
    if c == "5":
        return 5, 1
    return 1, 1


def parent_for(code: str, all_codes: list[str]) -> str | None:
    candidates = [c for c in all_codes if c != code and code.startswith(c)]
    return max(candidates, key=len) if candidates else None


def build_sql(accounts: list[dict]) -> str:
    codes = sorted({a["code"] for a in accounts}, key=lambda c: (len(c), c))
    code_set = set(codes)

    lines: list[str] = []
    lines.append("-- ══════════════════════════════════════════════════════")
    lines.append("-- Seed: شجرة الحسابات الكاملة (متزامن مع Accounts.xlsx)")
    lines.append("-- توليد تلقائي: python tools/sync_coa_from_xlsx.py")
    lines.append("-- يُصلح تلقائياً: تكرار 110103 الثاني → 110104 (العجز والزيادة)")
    lines.append("-- قم بتشغيل السكربت في قاعدة البيانات بعد أخذ نسخة احتياطية")
    lines.append("-- ══════════════════════════════════════════════════════")
    lines.append("")
    lines.append("SET FOREIGN_KEY_CHECKS = 0;")
    lines.append("DELETE FROM `JournalLines`;")
    lines.append("DELETE FROM `ReceiptVouchers`;")
    lines.append("DELETE FROM `PaymentVouchers`;")
    lines.append("DELETE FROM `JournalEntries`;")
    lines.append("DELETE FROM `Accounts`;")
    lines.append("")
    lines.append("ALTER TABLE `JournalLines` AUTO_INCREMENT = 1;")
    lines.append("ALTER TABLE `ReceiptVouchers` AUTO_INCREMENT = 1;")
    lines.append("ALTER TABLE `PaymentVouchers` AUTO_INCREMENT = 1;")
    lines.append("ALTER TABLE `JournalEntries` AUTO_INCREMENT = 1;")
    lines.append("ALTER TABLE `Accounts` AUTO_INCREMENT = 1;")
    lines.append("SET FOREIGN_KEY_CHECKS = 1;")
    lines.append("")
    lines.append("-- Type: 1=Asset, 2=Liability, 3=Equity, 4=Revenue, 5=Expense")
    lines.append("-- Nature: 1=Debit, 2=Credit — الربط بالأب يُشتق من أطول بادئة رقمية للكود")
    lines.append("")

    levels: dict[str, int] = {}

    def level_of(c: str) -> int:
        if c in levels:
            return levels[c]
        p = parent_for(c, list(code_set))
        if p is None:
            levels[c] = 1
        else:
            levels[c] = level_of(p) + 1
        return levels[c]

    for c in codes:
        level_of(c)

    by_section: dict[str, list[str]] = {"1": [], "2": [], "3": [], "4": [], "5": []}
    for c in codes:
        by_section[c[0]].append(c)

    section_title = {
        "1": "-- 1. الأصول",
        "2": "-- 2. الالتزامات",
        "3": "-- 3. حقوق الملكية",
        "4": "-- 4. الإيرادات",
        "5": "-- 5. المصاريف",
    }

    acc_map = {a["code"]: a for a in accounts}

    for sec in ("1", "2", "3", "4", "5"):
        sec_codes = sorted(by_section[sec], key=lambda c: (len(c), c))
        if not sec_codes:
            continue
        lines.append(section_title[sec])
        tuples: list[str] = []
        for code in sec_codes:
            a = acc_map[code]
            t, n = type_nature(code)
            p = parent_for(code, list(code_set))
            parent_sql = (
                "NULL"
                if p is None
                else f"(SELECT Id FROM (SELECT Id FROM Accounts WHERE Code='{p}') x)"
            )
            has_child = any(other != code and other.startswith(code) for other in code_set)
            is_leaf = 0 if has_child else 1
            lvl = levels[code]
            desc_sql = sql_str(a.get("desc"))
            name_sql = sql_str(a["name"])
            tuples.append(
                f"('{code}', {name_sql}, {desc_sql}, {t}, {n}, {parent_sql}, {lvl}, {is_leaf}, {a['allow']}, 1, NOW())"
            )
        lines.append("INSERT INTO `Accounts` (`Code`,`NameAr`,`Description`,`Type`,`Nature`,`ParentId`,`Level`,`IsLeaf`,`AllowPosting`,`IsSystem`,`CreatedAt`) VALUES")
        lines.append(",\n".join(tuples) + ";")
        lines.append("")

    return "\n".join(lines)


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else XLSX_DEFAULT
    if not xlsx.is_file():
        print(f"Missing file: {xlsx}", file=sys.stderr)
        sys.exit(1)
    accounts = load_accounts(xlsx)
    sql = build_sql(accounts)
    OUT_SQL.write_text(sql, encoding="utf-8")
    print(f"Wrote {OUT_SQL} ({len(accounts)} accounts)")


if __name__ == "__main__":
    main()
